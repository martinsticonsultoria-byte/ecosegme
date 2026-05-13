import os
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse, RedirectResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.field_sheet import FieldSheet
from app.models.sonus_upload import SonusUpload
from app.models.generated_report import GeneratedReport
from app.core.deps import get_current_user
from app.models.user import User
from app.pdf_generator import generate_laudo
from app.models.audit_log import AuditLog
from app.models.consolidated_report import ConsolidatedReport
from app import supabase_storage

router = APIRouter(prefix="/reports", tags=["reports"])

_MESES_PT = ["janeiro","fevereiro","março","abril","maio","junho",
             "julho","agosto","setembro","outubro","novembro","dezembro"]

def _fmt_sig_date(d):
    if not d:
        return ""
    return f"Manaus, {d.day:02d} de {_MESES_PT[d.month-1]} de {d.year}."

@router.post("/generate/{field_sheet_id}")
def generate_report(field_sheet_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    existing = db.query(GeneratedReport).filter(GeneratedReport.field_sheet_id == field_sheet_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Laudo ja gerado para esta ficha. Laudos sao imutaveis.")
    sheet = db.query(FieldSheet).filter(FieldSheet.id == field_sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Ficha nao encontrada")
    upload = db.query(SonusUpload).filter(SonusUpload.field_sheet_id == field_sheet_id).first()
    if not upload:
        raise HTTPException(status_code=400, detail="Nenhum PDF do SONUS 2 encontrado para esta ficha")
    employee = sheet.employee
    emp_nome = employee.nome if employee else (sheet.employee_name_text or "")
    emp_funcao = employee.funcao if employee else ""
    emp_matricula = employee.matricula if employee else ""
    emp_setor = employee.setor if employee else ""
    emp_local = employee.local if employee else ""

    data = {
        "laudo_number": str(sheet.laudo_number),
        "tipo_analise": sheet.tipo_analise or "Ruído Ocupacional",
        "razao_social": sheet.company.razao_social,
        "cnpj": sheet.company.cnpj or "",
        "endereco": sheet.company.endereco or "",
        "nome_funcionario": emp_nome,
        "matricula": emp_matricula,
        "funcao": emp_funcao,
        "setor": emp_setor,
        "local": emp_local,
        "turno": sheet.turno or "",
        "codigo_esocial": sheet.codigo_esocial or "",
        "collection_date": sheet.collection_date.strftime("%d/%m/%Y"),
        "technician_name": sheet.technician_name,
        "technician_name_2": sheet.technician_name_2 or "",
        "activity": sheet.activity or "",
        "machine_noise": sheet.machine_noise or "",
        "epi": sheet.epi or "",
        "pre_verificacao_db": sheet.pre_verificacao_db or "114,00",
        "pos_verificacao_db": sheet.pos_verificacao_db or "",
        "signature_date": sheet.signature_date.strftime("%d/%m/%Y") if sheet.signature_date else "",
        "parsed_employee_name": upload.parsed_employee_name or "",
        "inicio": upload.inicio or "",
        "fim": upload.fim or "",
        "dose_diaria": upload.dose_diaria or "",
        "ne_db": upload.ne_db or "",
        "nen_db": upload.nen_db or "",
        "tempo_medicao": upload.tempo_medicao or "",
    }
    try:
        output_path, filename, sha256 = generate_laudo(data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar PDF: {str(e)}")

    stored_path = output_path
    if supabase_storage.is_configured():
        try:
            with open(output_path, "rb") as f:
                pdf_bytes = f.read()
            supabase_storage.upload_pdf(pdf_bytes, filename)
            stored_path = f"supabase://{filename}"
            os.unlink(output_path)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao salvar PDF no storage: {str(e)}")

    report = GeneratedReport(
        field_sheet_id=field_sheet_id,
        sonus_upload_id=upload.id,
        output_path=stored_path,
        output_filename=filename,
        sha256_output=sha256,
        generated_by=current_user.id
    )
    db.add(report)
    db.flush()

    # Registra auditoria
    log = AuditLog(
        user_id=current_user.id,
        action="generate_report",
        field_sheet_id=field_sheet_id,
        original_filename=upload.original_filename,
        sha256_hash=sha256,
        details=f"Laudo gerado: {filename}"
    )
    db.add(log)
    db.commit()
    db.refresh(report)
    return {"id": report.id, "filename": filename, "sha256": sha256, "generated_at": report.generated_at, "download_url": f"/reports/download/{report.id}"}

@router.get("/download/{report_id}")
def download_report(report_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    report = db.query(GeneratedReport).filter(GeneratedReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Laudo nao encontrado")
    if report.output_path.startswith("supabase://"):
        storage_path = report.output_path.removeprefix("supabase://")
        try:
            signed_url = supabase_storage.get_signed_url(storage_path)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao gerar link de download: {str(e)}")
        return RedirectResponse(url=signed_url)
    if os.path.exists(report.output_path):
        return FileResponse(path=report.output_path, filename=report.output_filename, media_type="application/pdf")
    if supabase_storage.is_configured():
        try:
            signed_url = supabase_storage.get_signed_url(report.output_filename)
        except Exception:
            raise HTTPException(status_code=404, detail="Laudo nao encontrado no storage. Gere novamente.")
        return RedirectResponse(url=signed_url)
    raise HTTPException(status_code=404, detail="Arquivo do laudo nao encontrado. Gere novamente.")

@router.delete("/{report_id}")
def delete_report(report_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    report = db.query(GeneratedReport).filter(GeneratedReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Laudo não encontrado")
    if report.output_path.startswith("supabase://"):
        storage_path = report.output_path.removeprefix("supabase://")
        try:
            supabase_storage.delete_file(storage_path)
        except Exception:
            pass
    elif os.path.exists(report.output_path):
        try:
            os.unlink(report.output_path)
        except OSError:
            pass
    sheet = db.query(FieldSheet).filter(FieldSheet.id == report.field_sheet_id).first()
    db.delete(report)
    if sheet and sheet.status == "aprovada":
        sheet.status = "pendente"
        sheet.signature_date = None
    db.commit()
    return {"ok": True}


@router.get("/url/{report_id}")
def get_download_url(report_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    report = db.query(GeneratedReport).filter(GeneratedReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Laudo nao encontrado")
    if report.output_path.startswith("supabase://"):
        storage_path = report.output_path.removeprefix("supabase://")
        try:
            signed_url = supabase_storage.get_signed_url(storage_path)
            return {"url": signed_url, "filename": report.output_filename}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao gerar link: {str(e)}")
    if os.path.exists(report.output_path):
        return {"url": f"/reports/download/{report_id}", "filename": report.output_filename, "local": True}
    raise HTTPException(status_code=404, detail="Arquivo nao encontrado. Gere novamente.")

@router.get("/generate-bulk")
def generate_bulk_report(
    company_id: int,
    tipo_analise: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    import openpyxl
    import io
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    from app.models.sonus_upload import SonusUpload

    from sqlalchemy import or_
    tipo_filter = or_(FieldSheet.tipo_analise == tipo_analise, FieldSheet.tipo_analise.is_(None)) \
        if tipo_analise == "Ruído" else FieldSheet.tipo_analise == tipo_analise
    sheets = db.query(FieldSheet).filter(
        FieldSheet.company_id == company_id,
        tipo_filter
    ).order_by(FieldSheet.laudo_number).all()

    if not sheets:
        raise HTTPException(status_code=404, detail="Nenhuma ficha encontrada para esse grupo")

    sem_numero = [s for s in sheets if not s.laudo_number]
    if sem_numero:
        raise HTTPException(
            status_code=400,
            detail=f"{len(sem_numero)} ficha(s) sem Nº de Ordem definido. Defina todos antes de gerar o relatório."
        )

    uploads_list = db.query(SonusUpload).filter(
        SonusUpload.field_sheet_id.in_([s.id for s in sheets])
    ).all()
    uploads = {u.field_sheet_id: u for u in uploads_list}

    company = sheets[0].company
    dates = [s.collection_date for s in sheets]
    period = f"{min(dates).strftime('%d/%m/%Y')} à {max(dates).strftime('%d/%m/%Y')}"
    year = datetime.now().year

    template_path = os.path.join(os.path.dirname(__file__), "../templates/relatorio_template.xlsx")
    wb = openpyxl.load_workbook(template_path)

    # Capa
    capa = wb["Capa "]
    capa["M18"] = company.razao_social
    capa["M19"] = company.endereco or ""
    capa["M20"] = company.cnpj or ""
    capa["N23"] = period
    laudo_numbers = [s.laudo_number for s in sheets]
    capa["M3"] = f"{laudo_numbers[0]:04d}-1 ao {laudo_numbers[-1]:04d}-{len(sheets)}"
    # Preenche número das ordens na Capa e apaga placeholders do template
    all_number_cells = [
        "P3","P4","P5","P6","P7","P8","P9","P10","P11","P12",
        "S3","S4","S5","S6","S7","S8","S9","S10","S11","S12",
        "V3","V4","V5","V6","V7","V8","V9","V10","V11","V12",
        "Y3","Y4","Y5","Y6","Y7","Y8","Y9","Y10","Y11","Y12",
        "AB3","AB4","AB5","AB6","AB7","AB8","AB9","AB10","AB11","AB12",
    ]
    for i, coord in enumerate(all_number_cells):
        if i < len(sheets):
            capa[coord] = f"{sheets[i].laudo_number:04d}-{i+1}"
        else:
            capa[coord] = None  # limpa placeholder do template

    # Resumo
    resumo = wb["Resumo"]
    resumo["B1"] = company.razao_social
    resumo["F1"] = period
    resumo["B2"] = company.endereco or ""
    # Limpa todas as linhas de dados do template (5 a 54) antes de preencher
    for clear_row in range(5, 55):
        for col in range(1, 8):
            resumo.cell(row=clear_row, column=col).value = None
    for i, sheet in enumerate(sheets):
        row = 5 + i
        upload = uploads.get(sheet.id)
        emp = sheet.employee
        funcao = emp.funcao if emp else ""
        local_col = emp.local if emp else ""
        setor = emp.setor if emp else ""
        ne_val = None
        if upload and upload.ne_db:
            try:
                ne_val = float(upload.ne_db.replace(",", "."))
            except Exception:
                ne_val = None
        status = "N.C." if (ne_val is not None and ne_val > 85) else ("OK" if ne_val is not None else "—")
        resumo.cell(row=row, column=1, value=sheet.laudo_number)
        resumo.cell(row=row, column=2, value=funcao)
        resumo.cell(row=row, column=3, value=local_col)
        resumo.cell(row=row, column=4, value=setor)
        resumo.cell(row=row, column=5, value=ne_val)
        resumo.cell(row=row, column=6, value=85)
        resumo.cell(row=row, column=7, value=status)

    # Fichas individuais (clonar aba "1" para cada ficha)
    wb["1"].title = "_tpl"
    copies = [wb.copy_worksheet(wb["_tpl"]) for _ in sheets]
    for i, (sheet, ws) in enumerate(zip(sheets, copies)):
        ws.title = str(i + 1)
        upload = uploads.get(sheet.id)
        emp = sheet.employee
        emp_nome = emp.nome if emp else (sheet.employee_name_text or "")
        emp_funcao = emp.funcao if emp else ""
        emp_matricula = emp.matricula if emp else ""
        emp_setor = emp.setor if emp else ""
        emp_local = emp.local if emp else ""

        # Formata data de assinatura no padrão "Manaus, DD de MÊS de AAAA."
        _MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                  "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
        sig_date_str = ""
        if sheet.signature_date:
            sig_date_str = f"Manaus, {sheet.signature_date.day:02d} de {_MESES[sheet.signature_date.month-1]} de {sheet.signature_date.year}."

        ne_val_sheet = None
        if upload and upload.ne_db:
            try:
                ne_val_sheet = float(upload.ne_db.replace(",", "."))
            except Exception:
                ne_val_sheet = None

        conclusao_text = ""
        if ne_val_sheet is not None:
            if ne_val_sheet <= 85:
                conclusao_text = (
                    "O nível equivalente de ruído está dentro do limite de tolerância estabelecido "
                    "no Anexo 1 da NR 15 para uma jornada de 8 horas diárias, não caracterizando "
                    "exposição excessiva. No entanto, é recomendável manter medidas preventivas para "
                    "reduzir riscos auditivos e garantir a conformidade com as normas de segurança."
                )
            else:
                conclusao_text = (
                    "O nível equivalente de ruído excede o limite de tolerância do Anexo 1 da NR 15 "
                    "para uma jornada de 8 horas diárias, exigindo medidas preventivas imediatas. "
                    "Quando o ruído atinge ou supera o nível de ação, é fundamental minimizar os "
                    "riscos auditivos e garantir a conformidade com as normas de segurança."
                )
        acao_result = ""
        if ne_val_sheet is not None:
            acao_result = "ACIMA" if ne_val_sheet > 85 else "ABAIXO"

        ws["E1"] = f"{sheet.laudo_number:04d}-{i+1}"
        ws["L1"] = sig_date_str
        ws["B2"] = company.razao_social
        ws["B3"] = company.endereco or ""
        ws["B6"] = emp_nome
        ws["B7"] = emp_funcao
        ws["E7"] = emp_matricula
        ws["B8"] = emp_local
        ws["E8"] = emp_setor
        ws["B11"] = sheet.collection_date.strftime("%d/%m/%Y")
        ws["F11"] = sheet.technician_name
        ws["A14"] = sheet.activity or ""
        ws["A16"] = sheet.machine_noise or "Ausência de equipamentos ruidosos."
        ws["A18"] = sheet.epi or ""
        ws["C24"] = sheet.pre_verificacao_db or "114,00"
        ws["F24"] = sheet.pos_verificacao_db or ""
        if upload:
            ws["C29"] = upload.dose_diaria or ""
            ws["C30"] = upload.ne_db or ""
            ws["C31"] = upload.nen_db or ""
            ws["C32"] = upload.tempo_medicao or ""
            ws["I29"] = acao_result
        ws["A34"] = conclusao_text
        ws["A36"] = sig_date_str

    del wb["_tpl"]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    xlsx_bytes = output.read()

    safe_name = company.razao_social.replace(" ", "_")[:25]
    safe_tipo = tipo_analise.replace(" ", "_")
    filename = f"relatorio_{safe_name}_{safe_tipo}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    storage_path = filename
    if supabase_storage.is_configured():
        try:
            supabase_storage.upload_pdf(xlsx_bytes, filename)
            storage_path = f"supabase://{filename}"
        except Exception:
            storage_path = filename

    rec = ConsolidatedReport(
        company_id=company_id,
        tipo_analise=tipo_analise,
        format="xlsx",
        filename=filename,
        storage_path=storage_path,
        generated_by=current_user.id,
    )
    db.add(rec)
    db.commit()

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/generate-bulk-pdf")
def generate_bulk_pdf(
    company_id: int,
    tipo_analise: str,
    field_sheet_ids: Optional[List[int]] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    import io, tempfile
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    from jinja2 import Template
    from weasyprint import HTML
    from app.models.sonus_upload import SonusUpload
    from sqlalchemy import or_

    tipo_filter = or_(FieldSheet.tipo_analise == tipo_analise, FieldSheet.tipo_analise.is_(None)) \
        if tipo_analise == "Ruído" else FieldSheet.tipo_analise == tipo_analise
    if field_sheet_ids:
        sheets = db.query(FieldSheet).filter(
            FieldSheet.id.in_(field_sheet_ids),
            FieldSheet.company_id == company_id
        ).order_by(FieldSheet.laudo_number).all()
    else:
        sheets = db.query(FieldSheet).filter(
            FieldSheet.company_id == company_id,
            tipo_filter
        ).order_by(FieldSheet.laudo_number).all()

    if not sheets:
        raise HTTPException(status_code=404, detail="Nenhuma ficha encontrada para esse grupo")

    sem_numero = [s for s in sheets if not s.laudo_number]
    if sem_numero:
        raise HTTPException(
            status_code=400,
            detail=f"{len(sem_numero)} ficha(s) sem Nº de Ordem definido. Defina todos antes de gerar o relatório."
        )

    uploads_list = db.query(SonusUpload).filter(
        SonusUpload.field_sheet_id.in_([s.id for s in sheets])
    ).all()
    uploads = {u.field_sheet_id: u for u in uploads_list}

    company = sheets[0].company
    dates = [s.collection_date for s in sheets]
    period = f"{min(dates).strftime('%d/%m/%Y')} à {max(dates).strftime('%d/%m/%Y')}"

    generated_ids = {
        r.field_sheet_id
        for r in db.query(GeneratedReport.field_sheet_id).filter(
            GeneratedReport.field_sheet_id.in_([s.id for s in sheets])
        ).all()
    }

    fichas = []
    for sheet in sheets:
        upload = uploads.get(sheet.id)
        emp = sheet.employee
        ne_db = upload.ne_db if upload else None
        ne_val = None
        if ne_db:
            try:
                ne_val = float(ne_db.replace(",", "."))
            except Exception:
                ne_val = None
        sig_d = sheet.signature_date
        if sig_d:
            sig_date_ext = f"{sig_d.day:02d} de {_MESES_PT[sig_d.month-1]} de {sig_d.year}"
        else:
            _now = datetime.now()
            sig_date_ext = f"{_now.day:02d} de {_MESES_PT[_now.month-1]} de {_now.year}"
        fichas.append({
            "laudo_number": sheet.laudo_number,
            "employee_nome": emp.nome if emp else (sheet.employee_name_text or ""),
            "funcao": emp.funcao if emp else "",
            "matricula": emp.matricula if emp else "",
            "setor": emp.setor if emp else "",
            "local": emp.local if emp else "",
            "dosimeter_number": sheet.dosimeter_number,
            "collection_date": sheet.collection_date.strftime("%d/%m/%Y"),
            "technician_name": sheet.technician_name,
            "technician_name_2": sheet.technician_name_2 or "",
            "epi": sheet.epi or "",
            "activity": sheet.activity or "",
            "machine_noise": sheet.machine_noise or "",
            "pre_verificacao_db": sheet.pre_verificacao_db or "114,00",
            "pos_verificacao_db": sheet.pos_verificacao_db or "",
            "ne_db": ne_db or "",
            "ne_val": ne_val,
            "nen_db": upload.nen_db if upload else "",
            "dose_diaria": upload.dose_diaria if upload else "",
            "tempo_medicao": upload.tempo_medicao if upload else "",
            "inicio": upload.inicio if upload else "",
            "fim": upload.fim if upload else "",
            "signature_date": _fmt_sig_date(sheet.signature_date),
            "signature_date_ext": sig_date_ext,
            "has_laudo": sheet.id in generated_ids,
        })

    import base64
    tmpl_path = os.path.join(os.path.dirname(__file__), "../templates/relatorio_pdf.html")
    logo_path = os.path.join(os.path.dirname(__file__), "../templates/logo.png")
    assinatura_path = os.path.join(os.path.dirname(__file__), "../templates/relatório_assinatura.png")
    img_dir = os.path.join(os.path.dirname(__file__), "../templates/images")
    with open(logo_path, "rb") as f:
        logo_b64 = base64.b64encode(f.read()).decode()
    with open(assinatura_path, "rb") as f:
        assinatura_b64 = base64.b64encode(f.read()).decode()
    with open(os.path.join(img_dir, "capa_fundo.png.png"), "rb") as f:
        capa_fundo_b64 = base64.b64encode(f.read()).decode()
    with open(tmpl_path, "r", encoding="utf-8") as f:
        tmpl = Template(f.read())

    laudo_numbers = [s.laudo_number for s in sheets]
    html = tmpl.render(
        razao_social=company.razao_social,
        cnpj=company.cnpj or "",
        endereco=company.endereco or "",
        tipo_analise=tipo_analise,
        period=period,
        report_date=datetime.now().strftime("%m.%Y"),
        year=datetime.now().year,
        laudo_numbers=laudo_numbers,
        laudo_min=min(laudo_numbers) if laudo_numbers else '',
        laudo_max=max(laudo_numbers) if laudo_numbers else '',
        logo_b64=logo_b64,
        assinatura_b64=assinatura_b64,
        capa_fundo_b64=capa_fundo_b64,
        signature_date_ext=f"{datetime.now().day:02d} de {_MESES_PT[datetime.now().month-1]} de {datetime.now().year}",
        fichas=fichas,
    )

    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    HTML(string=html).write_pdf(tmp.name)
    tmp.close()

    safe_name = company.razao_social.replace(" ", "_")[:25]
    safe_tipo = tipo_analise.replace(" ", "_")
    filename = f"relatorio_{safe_name}_{safe_tipo}_{datetime.now().strftime('%Y%m%d')}.pdf"

    with open(tmp.name, "rb") as f:
        pdf_bytes = f.read()
    os.unlink(tmp.name)

    storage_path = filename
    if supabase_storage.is_configured():
        try:
            supabase_storage.upload_pdf(pdf_bytes, filename)
            storage_path = f"supabase://{filename}"
        except Exception:
            storage_path = filename

    rec = ConsolidatedReport(
        company_id=company_id,
        tipo_analise=tipo_analise,
        format="pdf",
        filename=filename,
        storage_path=storage_path,
        generated_by=current_user.id,
    )
    db.add(rec)
    db.commit()

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.delete("/consolidated/{rec_id}")
def delete_consolidated(rec_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    rec = db.query(ConsolidatedReport).filter(ConsolidatedReport.id == rec_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")
    if rec.storage_path.startswith("supabase://"):
        storage_path = rec.storage_path.removeprefix("supabase://")
        try:
            supabase_storage.delete_file(storage_path)
        except Exception:
            pass
    elif os.path.exists(rec.storage_path):
        try:
            os.unlink(rec.storage_path)
        except OSError:
            pass
    db.delete(rec)
    db.commit()
    return {"ok": True}


@router.get("/consolidated/{company_id}")
def list_consolidated(company_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    recs = db.query(ConsolidatedReport).filter(
        ConsolidatedReport.company_id == company_id
    ).order_by(ConsolidatedReport.generated_at.desc()).all()
    return [{"id": r.id, "filename": r.filename, "format": r.format, "tipo_analise": r.tipo_analise, "generated_at": r.generated_at} for r in recs]

@router.get("/consolidated/download/{rec_id}")
def download_consolidated(rec_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    rec = db.query(ConsolidatedReport).filter(ConsolidatedReport.id == rec_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")
    if rec.storage_path.startswith("supabase://"):
        storage_path = rec.storage_path.removeprefix("supabase://")
        try:
            signed_url = supabase_storage.get_signed_url(storage_path)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao gerar link: {str(e)}")
        return RedirectResponse(url=signed_url)
    if os.path.exists(rec.storage_path):
        media_type = "application/pdf" if rec.format == "pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return FileResponse(path=rec.storage_path, filename=rec.filename, media_type=media_type)
    raise HTTPException(status_code=404, detail="Arquivo não encontrado. Gere novamente.")

@router.get("/list/{company_id}")
def list_reports(company_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    reports = db.query(GeneratedReport).join(FieldSheet).filter(FieldSheet.company_id == company_id).order_by(GeneratedReport.generated_at.desc()).all()
    return [{"id": r.id, "filename": r.output_filename, "sha256": r.sha256_output, "generated_at": r.generated_at, "download_url": f"/reports/download/{r.id}"} for r in reports]
