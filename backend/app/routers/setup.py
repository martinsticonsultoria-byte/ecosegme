from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.models.user import User
from app.core.security import hash_password
from app.core.deps import require_admin

router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.post("/setup/seed")
def run_seed(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    if db.query(User).first():
        return {"message": "Setup já foi executado", "created": []}
    users = [
        {"name": "Admin EcoSegme", "email": "admin@ecosegme.com", "password": "Admin@2024", "role": "admin_staff"},
        {"name": "Técnico EcoSegme", "email": "tecnico@ecosegme.com", "password": "Tecnico@2024", "role": "technician"},
    ]
    created = []
    for u in users:
        exists = db.query(User).filter(User.email == u["email"]).first()
        if not exists:
            user = User(
                name=u["name"],
                email=u["email"],
                password_hash=hash_password(u["password"]),
                role=u["role"],
                active=True
            )
            db.add(user)
            created.append(u["email"])
    db.commit()
    return {"created": created, "message": "Seed executado com sucesso"}


# Planilha TLV (Google Sheets) — fonte viva do catálogo de agentes químicos.
# O sistema lê direto deste link a cada importação, então edições feitas na
# planilha (novos agentes, novos grupos, valores atualizados) são refletidas
# automaticamente na próxima vez que /setup/import-chemical-agents rodar.
TLV_SHEET_ID = "1KNKUBeDkb3VupB3FvUVsRn6HjXdPNM4J0kHwmw54qAI"
TLV_SHEET_EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{TLV_SHEET_ID}/export?format=xlsx"

# Cabeçalho da planilha (linha 1, aba "TLV") -> nome do campo em ChemicalAgent.
# Lido por NOME, não por posição — inserir/mover colunas na planilha não quebra
# a importação, só o texto do cabeçalho precisa continuar batendo.
_TLV_HEADER_MAP = {
    "Agente": "nome",
    "Substância / Agente": "_nome_fallback",  # usado só quando "Agente" vier vazio (maioria das linhas)
    "e-Social": "esocial",
    "Unidade": "unidade",
    "TWA": "acgih_twa",
    "STEL": "acgih_stel",
    "NR 15": "nr15_valor",
    "Bases de Efeitos Críticos": "efeito_critico",
    "Amostrador/Filtro": "amostrador",
    "Método": "metodo",
    "Método de Análise": "metodo_analise",
    "Vazão": "vazao",
    "Volume": "volume",
    "L.Q": "lq",
    "Resultados": "resultado_planilha",
    "CAS": "numero_cas",
    "Código CAS": "numero_cas",
    "Grupos": "grupo",
}


def _normalize_nome(nome):
    """Normaliza um nome de agente pra fins de comparação: unifica acentuação
    (NFC), colapsa espaços múltiplos e ignora maiúsculas/minúsculas. Evita que
    uma diferença mínima de texto entre edições da planilha (espaço extra,
    unicode diferente pro mesmo acento) crie um agente duplicado no catálogo
    em vez de atualizar o já existente."""
    import unicodedata
    if not nome:
        return ""
    nome = unicodedata.normalize("NFC", nome)
    return " ".join(nome.split()).casefold()


def _parse_tlv_xlsx(xlsx_file, sheet_name="TLV", header_row=1, data_start_row=2):
    """Lê a planilha TLV e devolve uma lista de dicts (um por agente), mapeando
    colunas pelo texto do cabeçalho (linha `header_row`). Não toca no banco —
    função pura, fácil de testar isoladamente. `xlsx_file` aceita tanto um
    caminho de arquivo quanto um objeto file-like (ex: BytesIO do download).
    """
    import openpyxl

    def _str(val):
        return str(val).strip() if val is not None else ""

    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    ws = wb[sheet_name]

    col_by_field = {}
    for cell in ws[header_row]:
        campo = _TLV_HEADER_MAP.get(_str(cell.value).rstrip())
        if campo and campo not in col_by_field:  # mantém a 1ª ocorrência (há "Agente" duplicado)
            col_by_field[campo] = cell.column

    if "nome" not in col_by_field and "_nome_fallback" not in col_by_field:
        raise ValueError(
            f'Nenhuma coluna de nome do agente ("Agente" ou "Substância / Agente") '
            f'encontrada no cabeçalho (linha {header_row}) de "{sheet_name}".'
        )

    agentes = []
    for row in range(data_start_row, ws.max_row + 1):
        # "Agente" só vem preenchido numa fração das linhas (as com dados
        # completos de TLV); a maioria só tem "Substância / Agente" — por
        # isso o nome cai pro fallback quando "Agente" está vazio.
        nome_principal = _str(ws.cell(row, col_by_field["nome"]).value) if "nome" in col_by_field else ""
        nome_fallback  = _str(ws.cell(row, col_by_field["_nome_fallback"]).value) if "_nome_fallback" in col_by_field else ""
        nome = nome_principal or nome_fallback
        if not nome:
            continue  # linha em branco

        dado = {"nome": nome}
        for campo, col in col_by_field.items():
            if campo in ("nome", "_nome_fallback"):
                continue
            valor = _str(ws.cell(row, col).value)
            dado[campo] = valor if (campo != "grupo" or valor) else None
        agentes.append(dado)

    return agentes


@router.post("/setup/import-chemical-agents")
def import_chemical_agents(
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    """
    Importa/atualiza o catálogo de agentes químicos a partir da planilha TLV
    (Google Sheets, TLV_SHEET_EXPORT_URL).

    IMPORTANTE: o mesmo agente (mesmo CAS/nome) pode legitimamente pertencer
    a mais de um grupo — ex.: Benzeno aparece em linhas separadas sem grupo,
    em "VOC", em "BETX (4)" e em "BETX (5)". Por isso a identidade de cada
    linha pro upsert é o par (CAS-ou-nome, grupo), nunca só o CAS/nome
    sozinho — do contrário, linhas do mesmo agente em grupos diferentes se
    confundiriam como duplicatas umas das outras.

    Passo 1: qualquer combinação (agente, grupo) que exista no banco mas não
    exista mais na planilha atual tem o grupo limpo (a planilha manda: se
    ela não lista mais aquele agente naquele grupo, o cadastro não deveria
    dizer que ele pertence lá).
    Passo 2: upsert normal de cada combinação (agente, grupo) que a planilha
    define — se houver mais de uma linha do banco pra mesma combinação
    (duplicata real), consolida numa só, exceto se alguma já estiver
    vinculada a uma ficha real (mantida intacta e apenas ignorada).
    """
    import io
    import httpx
    from sqlalchemy.exc import IntegrityError

    from app.models.chemical_agent import ChemicalAgent

    try:
        resp = httpx.get(TLV_SHEET_EXPORT_URL, follow_redirects=True, timeout=30.0)
        resp.raise_for_status()
    except httpx.HTTPError as e:
        raise HTTPException(
            status_code=502,
            detail=f"Erro ao baixar a planilha do Google Sheets: {e}",
        )

    try:
        agentes = _parse_tlv_xlsx(io.BytesIO(resp.content))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler planilha: {e}")

    def _identidade(cas, nome):
        cas = (cas or "").strip()
        return cas if cas else _normalize_nome(nome)

    def _grupo_key(grupo):
        grupo = (grupo or "").strip()
        return grupo or None

    # Uma linha por combinação (identidade, grupo) — se a planilha repetir a
    # mesma combinação em mais de uma linha, a última processada prevalece.
    sheet_por_chave = {}
    for dado in agentes:
        chave = (_identidade(dado.get("numero_cas"), dado["nome"]), _grupo_key(dado.get("grupo")))
        sheet_por_chave[chave] = dado

    def _indexar_banco():
        idx = {}
        for ag in db.query(ChemicalAgent).all():
            chave = (_identidade(ag.numero_cas, ag.nome), _grupo_key(ag.grupo))
            idx.setdefault(chave, []).append(ag)
        return idx

    cleaned = 0
    merged = 0
    inserted = 0
    updated = 0

    # Passo 1: limpa combinações (agente, grupo) que não existem mais na planilha
    for (identidade, grupo_key), linhas in _indexar_banco().items():
        if grupo_key is None or (identidade, grupo_key) in sheet_por_chave:
            continue
        for ag in linhas:
            ag.grupo = None
            cleaned += 1
    db.flush()

    # Passo 2: upsert de cada combinação (agente, grupo) que a planilha define,
    # reindexando o banco já sem as combinações obsoletas do passo 1
    db_por_chave = _indexar_banco()
    for chave, dado in sheet_por_chave.items():
        candidatos = db_por_chave.get(chave, [])
        if candidatos:
            principal, *duplicatas = candidatos
            for dup in duplicatas:
                try:
                    with db.begin_nested():
                        db.delete(dup)
                        db.flush()
                    merged += 1
                except IntegrityError:
                    pass  # ainda referenciado por alguma ficha — mantém como está
            for campo, valor in dado.items():
                setattr(principal, campo, valor)
            db.flush()
            updated += 1
        else:
            db.add(ChemicalAgent(**dado))
            db.flush()
            inserted += 1

    db.commit()
    return {
        "message": "Importação concluída",
        "inserted": inserted,
        "updated": updated,
        "merged": merged,
        "cleaned": cleaned,
        "total": inserted + updated,
    }
