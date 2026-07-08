from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date

from app.database import get_db
from app.models.chemical_field_sheet import ChemicalFieldSheet
from app.models.chemical_sheet_agent import ChemicalSheetAgent
from app.models.chemical_agent import ChemicalAgent
from app.models.employee import Employee
from app.schemas.chemical import (
    ChemicalFieldSheetCreate,
    ChemicalFieldSheetUpdate,
    ChemicalFieldSheetOut,
    ChemicalSheetAgentCreate,
    ChemicalSheetAgentUpdate,
    ChemicalSheetAgentOut,
)
from app.core.deps import get_current_user, require_admin
from app.models.user import User

router = APIRouter(prefix="/chemical-field-sheets", tags=["chemical-field-sheets"])


# ──────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────

def _calcular_resultado(valor: str, agent: ChemicalAgent) -> str:
    """Determina resultado_status a partir do valor medido e limites do agente."""
    if not valor or not valor.strip():
        return "pendente"
    v = valor.strip()
    if v.startswith("<"):
        return "nao_detectado"
    try:
        num = float(v.replace(",", "."))
        limite_str = agent.nr15_valor or agent.acgih_twa  # NR-15 tem prioridade; fallback ACGIH
        if not limite_str or limite_str.strip() in ("-", ""):
            return "pendente"
        limite = float(limite_str.replace(",", "."))
        return "dentro_limite" if num <= limite else "acima_limite"
    except (ValueError, TypeError):
        return "pendente"


def _get_sheet_or_404(sheet_id: int, db: Session) -> ChemicalFieldSheet:
    sheet = db.query(ChemicalFieldSheet).filter(ChemicalFieldSheet.id == sheet_id).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Ficha química não encontrada")
    return sheet


# ──────────────────────────────────────────────────────
# CRUD — Fichas
# ──────────────────────────────────────────────────────

@router.get("", response_model=List[ChemicalFieldSheetOut])
def list_chemical_field_sheets(
    company_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(ChemicalFieldSheet)
    if company_id:
        q = q.filter(ChemicalFieldSheet.company_id == company_id)
    if status:
        q = q.filter(ChemicalFieldSheet.status == status)
    return q.order_by(ChemicalFieldSheet.created_at.desc()).all()


@router.post("", response_model=ChemicalFieldSheetOut, status_code=201)
def create_chemical_field_sheet(
    data: ChemicalFieldSheetCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Resolve ou cria funcionário
    employee_id = data.employee_id
    if not employee_id and data.employee_name_text:
        existing = db.query(Employee).filter(
            Employee.company_id == data.company_id,
            Employee.nome == data.employee_name_text,
        ).first()
        if existing:
            employee_id = existing.id
        else:
            new_emp = Employee(
                company_id=data.company_id,
                nome=data.employee_name_text,
            )
            db.add(new_emp)
            db.flush()
            employee_id = new_emp.id

    from app.models.chemical_field_sheet import _CONCLUSAO_PADRAO
    sheet = ChemicalFieldSheet(
        **{k: v for k, v in data.dict().items() if k != "employee_id"},
        employee_id=employee_id,
        created_by=current_user.id,
        conclusao_texto=_CONCLUSAO_PADRAO,
    )
    db.add(sheet)
    db.commit()
    db.refresh(sheet)
    return sheet


@router.get("/report/pdf")
def generate_chemical_pdf_report(
    company_id: int,
    field_sheet_ids: Optional[List[int]] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Gera relatório PDF das fichas químicas de uma empresa (com capa e fichas individuais)."""
    import os, io, math, tempfile, re as _re, base64
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    from jinja2 import Template
    from weasyprint import HTML
    from app.models.company import Company
    from app.models.consolidated_report import ConsolidatedReport
    from app import supabase_storage

    _MESES_PT = ["janeiro","fevereiro","março","abril","maio","junho",
                 "julho","agosto","setembro","outubro","novembro","dezembro"]

    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")

    q = db.query(ChemicalFieldSheet).filter(ChemicalFieldSheet.company_id == company_id)
    if field_sheet_ids:
        q = q.filter(ChemicalFieldSheet.id.in_(field_sheet_ids))
    sheets = q.order_by(ChemicalFieldSheet.laudo_number, ChemicalFieldSheet.laudo_y).all()

    if not sheets:
        raise HTTPException(status_code=404, detail="Nenhuma ficha química encontrada para esta empresa")

    sem_numero = [s for s in sheets if not s.laudo_number]
    if sem_numero:
        raise HTTPException(
            status_code=400,
            detail=f"{len(sem_numero)} ficha(s) sem Nº do Laudo. Aprove todas as fichas antes de gerar o relatório."
        )

    year = datetime.now().year
    dates = [s.collection_date for s in sheets]
    period = f"{min(dates).strftime('%d/%m/%Y')} à {max(dates).strftime('%d/%m/%Y')}"

    # ── Prioridade de resultado: acima_limite > pendente > dentro_limite > nao_detectado ──
    _PRIO = {"acima_limite": 3, "pendente": 2, "dentro_limite": 1, "nao_detectado": 0}

    fichas = []
    for sheet in sheets:
        sig_d = sheet.data_relatorio or sheet.signature_date
        if sig_d:
            sig_date_ext = f"{sig_d.day:02d} de {_MESES_PT[sig_d.month-1]} de {sig_d.year}"
        else:
            _now = datetime.now()
            sig_date_ext = f"{_now.day:02d} de {_MESES_PT[_now.month-1]} de {_now.year}"

        agents_data = []
        pior_prio = -1
        resultado_geral = "pendente"
        for sa in (sheet.agents or []):
            ag = sa.agent
            rs = sa.resultado_status or "pendente"
            agents_data.append({
                "agent_nome":      ag.nome if ag else "—",
                "agent_cas":       ag.numero_cas if ag else "—",
                "agent_unidade":   ag.unidade if ag else "—",
                "agent_nr15":      ag.nr15_valor if ag else "—",
                "agent_acgih_twa": ag.acgih_twa if ag else "—",
                "agent_lq":        ag.lq if ag else "—",
                "valor_encontrado": sa.valor_encontrado or "—",
                "resultado_status": rs,
            })
            p = _PRIO.get(rs, 0)
            if p > pior_prio:
                pior_prio = p
                resultado_geral = rs
        if not agents_data:
            resultado_geral = "pendente"

        fichas.append({
            "laudo_number":       sheet.laudo_number,
            "laudo_y":            sheet.laudo_y or 1,
            "employee_nome":      (sheet.employee.nome if sheet.employee else sheet.employee_name_text) or "—",
            "funcao":             sheet.funcao or "—",
            "matricula":          sheet.matricula or "—",
            "setor":              sheet.setor or "—",
            "local":              sheet.local or "—",
            "collection_date":    sheet.collection_date.strftime("%d/%m/%Y"),
            "technician_name":    sheet.technician_name or "—",
            "numero_amostrador":  sheet.numero_amostrador or "—",
            "tipo_amostrador":    sheet.tipo_amostrador or "—",
            "situacao_ambiente":  sheet.situacao_ambiente or "",
            "atividade":          sheet.atividade or "",
            "jornada_trabalho":   sheet.jornada_trabalho or "",
            "volume_ar_amostrado": sheet.volume_ar_amostrado or "",
            "epi":                sheet.epi or "",
            "observacoes":        sheet.observacoes or "",
            "conclusao_texto":    sheet.conclusao_texto or "",
            "signature_date_ext": sig_date_ext,
            "agents":             agents_data,
            "resultado_geral":    resultado_geral,
        })

    # ── Calcula fontes e range de laudos (mesma lógica do relatório de ruído) ──
    def calc_font_size(texto, max_width_pt, font_size_pt=16.5, min_pt=9.0, char_factor=0.55):
        while font_size_pt > min_pt:
            if math.ceil(len(texto) / (max_width_pt / (char_factor * font_size_pt))) <= 2:
                break
            font_size_pt -= 0.5
        return f'{font_size_pt}pt'

    empresa_font_size  = calc_font_size(company.razao_social or '', 375.9)
    endereco_font_size = calc_font_size(company.endereco or '', 329.7)

    sheets_com_y = [s for s in sheets if s.laudo_y is not None]
    if sheets_com_y:
        sorted_y = sorted(sheets_com_y, key=lambda s: (s.laudo_number or '', s.laudo_y))
        laudo_min = f"{sorted_y[0].laudo_number}.{sorted_y[0].laudo_y}"
        laudo_max = f"{sorted_y[-1].laudo_number}.{sorted_y[-1].laudo_y}"
    else:
        sorted_all = sorted(sheets, key=lambda s: s.laudo_number or '')
        laudo_min = sorted_all[0].laudo_number or 'SN'
        laudo_max = sorted_all[-1].laudo_number or 'SN'

    nr_texto     = f"{laudo_min}/{year} ao {laudo_max}/{year}" if laudo_min != laudo_max else f"{laudo_min}/{year}"
    nr_font_size = calc_font_size(nr_texto, 321.1, font_size_pt=20.0, min_pt=10.0, char_factor=0.65)

    _rel_dates = [s.data_relatorio for s in sheets if s.data_relatorio]
    _ref_date  = max(_rel_dates) if _rel_dates else datetime.now().date()
    signature_date_ext = f"{_ref_date.day:02d} de {_MESES_PT[_ref_date.month-1]} de {_ref_date.year}"

    # ── Carrega imagens e template ────────────────────────────────────────────
    tmpl_dir   = os.path.join(os.path.dirname(__file__), "../templates")
    logo_path  = os.path.join(tmpl_dir, "logo.png")
    assin_path = os.path.join(tmpl_dir, "relatório_assinatura.png")
    fundo_path = os.path.join(tmpl_dir, "images", "capa_fundo.png.png")
    tmpl_path  = os.path.join(tmpl_dir, "relatorio_quimico_pdf.html")

    with open(logo_path,  "rb") as f: logo_b64       = base64.b64encode(f.read()).decode()
    with open(assin_path, "rb") as f: assinatura_b64 = base64.b64encode(f.read()).decode()
    with open(fundo_path, "rb") as f: capa_fundo_b64 = base64.b64encode(f.read()).decode()
    with open(tmpl_path,  "r", encoding="utf-8") as f: tmpl = Template(f.read())

    html = tmpl.render(
        razao_social=company.razao_social,
        cnpj=company.cnpj or "",
        endereco=company.endereco or "",
        period=period,
        report_date=datetime.now().strftime("%m.%Y"),
        year=year,
        laudo_min=laudo_min,
        laudo_max=laudo_max,
        logo_b64=logo_b64,
        assinatura_b64=assinatura_b64,
        capa_fundo_b64=capa_fundo_b64,
        empresa_font_size=empresa_font_size,
        endereco_font_size=endereco_font_size,
        nr_font_size=nr_font_size,
        signature_date_ext=signature_date_ext,
        fichas=fichas,
    )

    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    HTML(string=html).write_pdf(tmp.name)
    tmp.close()
    with open(tmp.name, "rb") as f: pdf_bytes = f.read()
    os.unlink(tmp.name)

    safe_name = _re.sub(r'[\\/:*?"<>|\s]+', '_', company.razao_social or 'Empresa').strip('_')[:20]
    first_num  = sorted_y[0].laudo_number if sheets_com_y else (sheets[0].laudo_number or 'SN')
    filename   = f"Relatório_Químico_{safe_name}_{first_num}.pdf"

    storage_path = filename
    if supabase_storage.is_configured():
        try:
            supabase_storage.upload_pdf(pdf_bytes, filename)
            storage_path = f"supabase://{filename}"
        except Exception:
            pass

    rec = ConsolidatedReport(
        company_id=company_id,
        tipo_analise="Químico",
        format="pdf",
        filename=filename,
        storage_path=storage_path,
        generated_by=current_user.id,
    )
    db.add(rec)
    db.commit()

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/report/xlsx")
def generate_chemical_xlsx_report(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Gera relatório XLSX consolidado das fichas químicas de uma empresa."""
    import io, re as _re
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    from app.models.company import Company
    from app.models.consolidated_report import ConsolidatedReport
    from app import supabase_storage

    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")

    sheets = (
        db.query(ChemicalFieldSheet)
        .filter(ChemicalFieldSheet.company_id == company_id)
        .order_by(ChemicalFieldSheet.laudo_number, ChemicalFieldSheet.collection_date)
        .all()
    )
    if not sheets:
        raise HTTPException(status_code=404, detail="Nenhuma ficha química encontrada para esta empresa")

    year = datetime.now().year

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agentes Químicos"

    # ── Estilos ──────────────────────────────────────────────
    _green      = "1A7A3C"
    header_fill = PatternFill(start_color=_green, end_color=_green, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center_al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    STATUS_LABEL = {
        "dentro_limite": "Dentro do Limite",
        "acima_limite":  "Acima do Limite",
        "nao_detectado": "Não Detectado",
        "pendente":      "Pendente",
    }
    STATUS_FILL = {
        "dentro_limite": ("D1FAE5", "166534"),
        "acima_limite":  ("FEE2E2", "991B1B"),
        "nao_detectado": ("DBEAFE", "1E40AF"),
        "pendente":      ("FEF9C3", "854D0E"),
    }

    # ── Cabeçalho do documento ────────────────────────────────
    NUM_COLS = 15
    last_col = get_column_letter(NUM_COLS)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "RELATÓRIO DE MONITORAMENTO DE AGENTES QUÍMICOS"
    ws["A1"].font = Font(bold=True, size=14, color=_green)
    ws["A1"].alignment = center_al

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = company.razao_social
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = center_al

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = f"CNPJ: {company.cnpj or '—'}   |   Endereço: {company.endereco or '—'}"
    ws["A3"].alignment = center_al
    ws["A3"].font = Font(size=10, color="475569")

    ws.merge_cells(f"A4:{last_col}4")
    ws["A4"] = f"Emitido em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A4"].alignment = center_al
    ws["A4"].font = Font(size=9, color="94A3B8")

    ws.append([])  # linha em branco (linha 5)

    # ── Cabeçalho da tabela (linha 6) ─────────────────────────
    headers = [
        "Nº Laudo", "Funcionário", "Função", "Setor", "Local",
        "Data Coleta", "Amostrador", "Tipo Amostrador",
        "Agente Químico", "Nº CAS", "Unidade",
        "Valor Encontrado", "LT NR-15", "TLV-TWA ACGIH",
        "Resultado",
    ]
    ws.append(headers)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_al
        cell.border = thin_border
    ws.row_dimensions[hdr_row].height = 32

    # ── Linhas de dados ───────────────────────────────────────
    for sheet in sheets:
        laudo_str = (
            f"{sheet.laudo_number}.{sheet.laudo_y or 1}/{year}"
            if sheet.laudo_number else "S/Nº"
        )
        date_str = sheet.collection_date.strftime("%d/%m/%Y") if sheet.collection_date else ""
        emp_nome = (sheet.employee.nome if sheet.employee else sheet.employee_name_text) or "—"
        agents   = sheet.agents or []

        def write_row(extra_cols, _sheet=sheet, _laudo=laudo_str, _emp=emp_nome, _date=date_str):
            row_data = [
                _laudo, _emp, _sheet.funcao or "—", _sheet.setor or "—", _sheet.local or "—",
                _date, _sheet.numero_amostrador or "—", _sheet.tipo_amostrador or "—",
            ] + extra_cols
            ws.append(row_data)
            r = ws.max_row
            for cell in ws[r]:
                cell.alignment = left_al
                cell.border = thin_border
            return r

        if not agents:
            write_row(["—", "—", "—", "—", "—", "—", "—"])
        else:
            for sa in agents:
                ag = sa.agent
                result_key = sa.resultado_status or "pendente"
                row_idx = write_row([
                    ag.nome if ag else "—",
                    ag.numero_cas if ag else "—",
                    ag.unidade if ag else "—",
                    sa.valor_encontrado or "—",
                    ag.nr15_valor if ag else "—",
                    ag.acgih_twa if ag else "—",
                    STATUS_LABEL.get(result_key, result_key),
                ])
                fill_color, font_color = STATUS_FILL.get(result_key, ("FFFFFF", "000000"))
                rc = ws.cell(row=row_idx, column=15)
                rc.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                rc.font = Font(bold=True, color=font_color, size=10)

    # ── Larguras das colunas ──────────────────────────────────
    col_widths = [14, 24, 18, 16, 16, 12, 20, 20, 32, 14, 10, 16, 12, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A7"  # congela cabeçalhos ao rolar

    # ── Serializa e salva ──────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    xlsx_bytes = output.read()

    safe_name = _re.sub(r'[\\/:*?"<>|\s]+', '_', company.razao_social or 'Empresa').strip('_')[:20]
    filename   = f"Relatório_Químico_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    storage_path = filename
    if supabase_storage.is_configured():
        try:
            supabase_storage.upload_pdf(xlsx_bytes, filename)
            storage_path = f"supabase://{filename}"
        except Exception:
            pass

    rec = ConsolidatedReport(
        company_id=company_id,
        tipo_analise="Químico",
        format="xlsx",
        filename=filename,
        storage_path=storage_path,
        generated_by=current_user.id,
    )
    db.add(rec)
    db.commit()

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{sheet_id}", response_model=ChemicalFieldSheetOut)
def get_chemical_field_sheet(
    sheet_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return _get_sheet_or_404(sheet_id, db)


@router.patch("/{sheet_id}", response_model=ChemicalFieldSheetOut)
def update_chemical_field_sheet(
    sheet_id: int,
    data: ChemicalFieldSheetUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    sheet = _get_sheet_or_404(sheet_id, db)
    for field, value in data.dict(exclude_unset=True).items():
        setattr(sheet, field, value)
    db.commit()
    db.refresh(sheet)
    return sheet


@router.patch("/{sheet_id}/status", response_model=ChemicalFieldSheetOut)
def approve_chemical_field_sheet(
    sheet_id: int,
    laudo_number: str = Query(..., description="Número do laudo (apenas o inteiro)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Aprova a ficha: define laudo_number, laudo_y, status e signature_date.
    laudo_y é calculado como count+1 de fichas já aprovadas com o mesmo laudo_number
    na mesma empresa — idêntico à regra das fichas de ruído."""
    sheet = _get_sheet_or_404(sheet_id, db)
    # Conta fichas já aprovadas com mesmo laudo_number na mesma empresa (exceto a atual)
    count = db.query(ChemicalFieldSheet).filter(
        ChemicalFieldSheet.company_id == sheet.company_id,
        ChemicalFieldSheet.laudo_number == laudo_number,
        ChemicalFieldSheet.laudo_y.isnot(None),
        ChemicalFieldSheet.id != sheet_id,
    ).count()
    sheet.laudo_number   = laudo_number
    sheet.laudo_y        = count + 1
    sheet.status         = "aprovado"
    sheet.signature_date = date.today()
    sheet.data_relatorio = date.today()
    db.commit()
    db.refresh(sheet)
    return sheet


@router.delete("/{sheet_id}", status_code=204)
def delete_chemical_field_sheet(
    sheet_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    sheet = _get_sheet_or_404(sheet_id, db)
    db.delete(sheet)
    db.commit()


# ──────────────────────────────────────────────────────
# Agentes vinculados à ficha
# ──────────────────────────────────────────────────────

@router.get("/{sheet_id}/agents", response_model=List[ChemicalSheetAgentOut])
def list_sheet_agents(
    sheet_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    _get_sheet_or_404(sheet_id, db)
    return (
        db.query(ChemicalSheetAgent)
        .filter(ChemicalSheetAgent.chemical_sheet_id == sheet_id)
        .order_by(ChemicalSheetAgent.id)
        .all()
    )


@router.post("/{sheet_id}/agents", response_model=ChemicalSheetAgentOut, status_code=201)
def add_sheet_agent(
    sheet_id: int,
    data: ChemicalSheetAgentCreate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    _get_sheet_or_404(sheet_id, db)
    agent = db.query(ChemicalAgent).filter(ChemicalAgent.id == data.agent_id).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agente não encontrado no catálogo")

    resultado = _calcular_resultado(data.valor_encontrado or "", agent)

    sa = ChemicalSheetAgent(
        chemical_sheet_id=sheet_id,
        agent_id=data.agent_id,
        valor_encontrado=data.valor_encontrado,
        resultado_status=data.resultado_status or resultado,
        observacao=data.observacao,
    )
    db.add(sa)
    db.commit()
    db.refresh(sa)
    return sa


@router.patch("/{sheet_id}/agents/{agent_id}", response_model=ChemicalSheetAgentOut)
def update_sheet_agent(
    sheet_id: int,
    agent_id: int,
    data: ChemicalSheetAgentUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    sa = (
        db.query(ChemicalSheetAgent)
        .filter(
            ChemicalSheetAgent.chemical_sheet_id == sheet_id,
            ChemicalSheetAgent.agent_id == agent_id,
        )
        .first()
    )
    if not sa:
        raise HTTPException(status_code=404, detail="Vínculo agente/ficha não encontrado")

    update_data = data.dict(exclude_unset=True)

    # Recalcula resultado_status automaticamente se valor_encontrado foi atualizado
    if "valor_encontrado" in update_data and "resultado_status" not in update_data:
        agent = db.query(ChemicalAgent).filter(ChemicalAgent.id == agent_id).first()
        update_data["resultado_status"] = _calcular_resultado(
            update_data["valor_encontrado"] or "", agent
        )

    for field, value in update_data.items():
        setattr(sa, field, value)

    db.commit()
    db.refresh(sa)
    return sa


@router.delete("/{sheet_id}/agents/{agent_id}", status_code=204)
def remove_sheet_agent(
    sheet_id: int,
    agent_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    sa = (
        db.query(ChemicalSheetAgent)
        .filter(
            ChemicalSheetAgent.chemical_sheet_id == sheet_id,
            ChemicalSheetAgent.agent_id == agent_id,
        )
        .first()
    )
    if not sa:
        raise HTTPException(status_code=404, detail="Vínculo agente/ficha não encontrado")
    db.delete(sa)
    db.commit()
