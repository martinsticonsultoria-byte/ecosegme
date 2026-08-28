from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
import io
import openpyxl
from app.database import get_db
from app.models.company import Company
from app.schemas.company import CompanyCreate, CompanyOut
from app.core.deps import get_current_user, require_admin

router = APIRouter(prefix="/companies", tags=["companies"])

@router.get("", response_model=List[CompanyOut])
def list_companies(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return db.query(Company).order_by(Company.razao_social).all()

@router.get("/bulk-template")
def download_bulk_template(_=Depends(get_current_user)):
    """Baixa planilha modelo para importação em massa de empresas."""
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empresas"
    ws.append(["Razão Social", "CNPJ", "Endereço"])
    ws.append(["Nome da Empresa Ltda", "00.000.000/0001-00", "Rua Exemplo, 123 - Manaus/AM"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_empresas.xlsx"}
    )

@router.post("/bulk-upload")
async def bulk_upload_companies(file: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(require_admin)):
    """
    Lê cabeçalhos da linha 1 e mapeia por nome (case-insensitive).
    Colunas reconhecidas: Razão Social/Empresa/Nome, CNPJ, Endereço
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo inválido. Use .xlsx")

    headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h:
                    return i
        return None

    idx_razao   = col(['razão social', 'razao social', 'empresa', 'nome'])
    idx_cnpj    = col(['cnpj'])
    idx_endereco= col(['endereço', 'endereco'])

    if idx_razao is None:
        raise HTTPException(status_code=400, detail="Planilha deve ter a coluna 'Razão Social' (ou 'Empresa')")

    def cell(row, idx):
        if idx is None or idx >= len(row) or row[idx] is None:
            return ''
        return str(row[idx]).strip()

    created = 0
    skipped = 0
    errors = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        razao_social = cell(row, idx_razao)
        cnpj = cell(row, idx_cnpj) or None
        endereco = cell(row, idx_endereco) or None

        if not razao_social or razao_social.lower() in ('none', 'nan', ''):
            skipped += 1
            continue

        existing = None
        if cnpj:
            existing = db.query(Company).filter(Company.cnpj == cnpj).first()
        if not existing:
            existing = db.query(Company).filter(Company.razao_social == razao_social).first()
        if existing:
            skipped += 1
            errors.append(f"Linha {i}: empresa '{razao_social}' já cadastrada — ignorada")
            continue

        db.add(Company(razao_social=razao_social, cnpj=cnpj, endereco=endereco))
        created += 1

    db.commit()
    return {"criados": created, "ignorados": skipped, "erros": errors}

@router.post("", response_model=CompanyOut)
def create_company(data: CompanyCreate, db: Session = Depends(get_db), _=Depends(require_admin)):
    company = Company(**data.dict())
    db.add(company)
    db.commit()
    db.refresh(company)
    return company

@router.get("/{company_id}", response_model=CompanyOut)
def get_company(company_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    return company

@router.put("/{company_id}", response_model=CompanyOut)
def update_company(company_id: int, data: CompanyCreate, db: Session = Depends(get_db), _=Depends(require_admin)):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    for key, value in data.dict().items():
        setattr(company, key, value)
    db.commit()
    db.refresh(company)
    return company

@router.delete("/{company_id}")
def delete_company(company_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    import os
    from app.models.generated_report import GeneratedReport
    from app.models.sonus_upload import SonusUpload
    from app.models.audit_log import AuditLog
    from app.models.field_sheet import FieldSheet
    from app.models.employee import Employee
    from app.models.consolidated_report import ConsolidatedReport
    from app import supabase_storage

    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")

    sheets = db.query(FieldSheet).filter(FieldSheet.company_id == company_id).all()
    sheet_ids = [s.id for s in sheets]

    if sheet_ids:
        # 1. Nullify AuditLogs vinculados às fichas
        db.query(AuditLog).filter(AuditLog.field_sheet_id.in_(sheet_ids)).update(
            {AuditLog.field_sheet_id: None}, synchronize_session=False
        )
        # 2. Delete GeneratedReports + arquivos no storage
        reports = db.query(GeneratedReport).filter(
            GeneratedReport.field_sheet_id.in_(sheet_ids)
        ).all()
        for report in reports:
            if report.output_path and report.output_path.startswith("supabase://"):
                try:
                    supabase_storage.delete_file(report.output_path.removeprefix("supabase://"))
                except Exception:
                    pass
            elif report.output_path and os.path.exists(report.output_path):
                try:
                    os.unlink(report.output_path)
                except OSError:
                    pass
            db.delete(report)
        db.flush()
        # 3. Delete SonusUploads
        db.query(SonusUpload).filter(
            SonusUpload.field_sheet_id.in_(sheet_ids)
        ).delete(synchronize_session=False)
        # 4. Delete FieldSheets
        db.query(FieldSheet).filter(
            FieldSheet.company_id == company_id
        ).delete(synchronize_session=False)

    # 5. Delete Employees
    db.query(Employee).filter(
        Employee.company_id == company_id
    ).delete(synchronize_session=False)

    # 6. Delete ConsolidatedReports + arquivos no storage
    consolidated = db.query(ConsolidatedReport).filter(
        ConsolidatedReport.company_id == company_id
    ).all()
    for rec in consolidated:
        if rec.storage_path and rec.storage_path.startswith("supabase://"):
            try:
                supabase_storage.delete_file(rec.storage_path.removeprefix("supabase://"))
            except Exception:
                pass
        elif rec.storage_path and os.path.exists(rec.storage_path):
            try:
                os.unlink(rec.storage_path)
            except OSError:
                pass
        db.delete(rec)
    db.flush()

    db.delete(company)
    db.commit()
    return {"ok": True}
