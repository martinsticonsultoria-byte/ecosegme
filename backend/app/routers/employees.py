from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
import io
import openpyxl
from app.database import get_db
from app.models.employee import Employee
from app.models.company import Company
from app.schemas.employee import EmployeeCreate, EmployeeOut
from app.core.deps import get_current_user, require_admin

router = APIRouter(prefix="/employees", tags=["employees"])

@router.get("", response_model=List[EmployeeOut])
def list_employees(company_id: int = None, db: Session = Depends(get_db), _=Depends(get_current_user)):
    query = db.query(Employee)
    if company_id:
        query = query.filter(Employee.company_id == company_id)
    return query.order_by(Employee.nome).all()

@router.post("", response_model=EmployeeOut)
def create_employee(data: EmployeeCreate, db: Session = Depends(get_db), _=Depends(get_current_user)):
    employee = Employee(**data.dict())
    db.add(employee)
    db.commit()
    db.refresh(employee)
    return employee

@router.put("/{employee_id}", response_model=EmployeeOut)
def update_employee(employee_id: int, data: EmployeeCreate, db: Session = Depends(get_db), _=Depends(require_admin)):
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    for key, value in data.dict().items():
        setattr(employee, key, value)
    db.commit()
    db.refresh(employee)
    return employee

@router.delete("/{employee_id}")
def delete_employee(employee_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    from app.models.field_sheet import FieldSheet
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")
    has_sheets = db.query(FieldSheet).filter(FieldSheet.employee_id == employee_id).first()
    if has_sheets:
        raise HTTPException(status_code=409, detail="Não é possível excluir este funcionário pois ele possui fichas de campo vinculadas.")
    db.delete(employee)
    db.commit()
    return {"ok": True}

@router.get("/bulk-template")
def download_bulk_template(_=Depends(get_current_user)):
    """Baixa planilha modelo para importação em massa de funcionários."""
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Funcionários"
    ws.append(["Empresa", "Nome", "Função", "Matrícula", "Setor", "Local"])
    ws.append(["Nome da Empresa Ltda", "João da Silva", "Operador", "001", "Produção", "Galpão A"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_funcionarios.xlsx"}
    )

@router.post("/bulk-upload")
async def bulk_upload_employees(file: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(require_admin)):
    """
    Lê cabeçalhos da linha 1 e mapeia por nome (case-insensitive).
    Colunas reconhecidas: Empresa, Nome/Funcionário, Função/Cargo, Matrícula, Setor, Local
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo inválido. Use .xlsx")

    # Mapeia cabeçalhos por nome
    headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h:
                    return i
        return None

    idx_empresa  = col(['empresa', 'company'])
    idx_nome     = col(['funcionário', 'funcionario', 'nome', 'colaborador'])
    idx_funcao   = col(['função', 'funcao', 'cargo'])
    idx_matricula= col(['matrícula', 'matricula', 'identificador'])
    idx_setor    = col(['setor'])
    idx_local    = col(['local'])

    if idx_empresa is None or idx_nome is None:
        raise HTTPException(status_code=400, detail="Planilha deve ter colunas 'Empresa' e 'Nome' (ou 'Funcionário')")

    def cell(row, idx):
        if idx is None or idx >= len(row) or row[idx] is None:
            return ''
        return str(row[idx]).strip()

    created = 0
    skipped = 0
    errors = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        empresa_val = cell(row, idx_empresa)
        nome = cell(row, idx_nome)

        if not nome or nome.lower() in ('none', 'nan', ''):
            skipped += 1
            continue

        company = None
        if empresa_val.isdigit():
            company = db.query(Company).filter(Company.id == int(empresa_val)).first()
        elif empresa_val:
            company = db.query(Company).filter(
                Company.razao_social.ilike(f'%{empresa_val}%')
            ).first()

        if not company:
            errors.append(f"Linha {i}: empresa '{empresa_val}' não encontrada")
            skipped += 1
            continue

        existing = db.query(Employee).filter(
            Employee.company_id == company.id,
            Employee.nome == nome
        ).first()
        if existing:
            skipped += 1
            continue

        emp = Employee(
            company_id=company.id,
            nome=nome,
            funcao=cell(row, idx_funcao) or None,
            matricula=cell(row, idx_matricula) or None,
            setor=cell(row, idx_setor) or None,
            local=cell(row, idx_local) or None,
        )
        db.add(emp)
        created += 1

    db.commit()
    return {"criados": created, "ignorados": skipped, "erros": errors}
